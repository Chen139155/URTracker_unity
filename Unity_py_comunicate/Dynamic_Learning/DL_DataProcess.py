import torch
import torch.nn.functional as F
from openpyxl import load_workbook
import math
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
import util
import copy
import numpy as np
import time
import os

"""
Function:
    数据预处理模块
Author:
    MuFengjun, CC 
"""


def distance_between_two_points(pointA, pointB):
    """计算两点之间距离"""
    x1, y1 = pointA
    x2, y2 = pointB
    return math.sqrt((x2 - x1) ** 2 + (y2 - y1) ** 2)


def get_urtracker_data(path="output/trackballdata.xlsx", type="mode1", Debug = False):
    """从本地获取urtracker数据并进行预处理"""
    ''' 配色 '''
    colors = ['#F79E7B','#9BABD2','#E48DBC','#BFDC84','#BC6F45'] # 低饱和 橙，烟蓝，粉，绿，棕
    ''' 创建存储路径文件夹'''
    date_str = time.strftime("%m%d_%H%M")
    path_figsave = os.path.join("output/fig/",date_str)
    isExists = os.path.exists(path_figsave)
    if not isExists:
        os.makedirs(path_figsave)
    else:
        print("目录已存在")

    # tobii on surface
    center_point = torch.tensor([540, 540])  # 用作归一化
    radius_point = 360

    workbook = load_workbook(filename=path)
    sheet = workbook.active

    # 读取数据
    raw_data = {}  # 该字典的key值为excel第一行的标题，value值为excel每一列的列表
    keys = []
    """
    {'time':[ ], 'target_x':[], 'target_y', 'cursor_x', 'cursor_y', 'Hex_x', 'Hex_y', 'Hex_z', 
    'linear_x', 'linear_y', 'linear_z', 'pose_x', 'pose_y', 'pose_z', 'pose_rx', 'pose_ry', 'pose_rz', 'Gaze_x', 'Gaze_y', 'stage_id', 'congnitive_flag', 'motor_flag'}
    """
    for row in sheet.iter_rows(values_only=True):
        if len(raw_data.keys()) == 0:
            for key in row:
                if key is not None:
                    keys.append(key)
                    raw_data[key] = []
            continue
        for i in range(len(keys)):
            raw_data[keys[i]].append(row[i + 1])

    raw_data_ori = copy.deepcopy(raw_data)
    raw_data_len = len(raw_data["time"])
    target_angel = torch.zeros(raw_data_len)
    ur_track_angel = torch.zeros(raw_data_len)
    gaze_track_angel = torch.zeros(raw_data_len)
    track_noise = torch.zeros(raw_data_len)

    raw_x = torch.zeros(raw_data_len)
    raw_y = torch.zeros(raw_data_len)

    raw_data["ori_x"] = raw_data["cursor_x"]
    raw_data["ori_y"] = raw_data["cursor_y"]
    raw_data["ori_gaze_x"] = raw_data["Gaze_x"]
    raw_data["ori_gaze_y"] = raw_data["Gaze_y"]
    
    ''' 获取静息态数据长度 '''
    time_start = raw_data["time"][0]
    time_track_start = 0.0
    raw_data_rest_len = 0
    for i in range(raw_data_len):
        time_track_start = raw_data["time"][i]
        if time_track_start - time_start >= 20:
            raw_data_rest_len = i+1
            break

    if Debug:
        # UR光标原始数据时序图
        plt.plot(raw_data["ori_x"], label="UR_x")
        plt.xlabel("Frame")
        plt.plot(raw_data["ori_y"], label="UR_y")
        plt.plot(raw_data["ori_gaze_x"], label = "Gaze_x")
        plt.plot(raw_data["ori_gaze_y"], label = "Gaze_y")
        plt.ylabel("Value")

        plt.savefig(os.path.join(path_figsave,"ori_x_y"+type))
        # plt.xlim(220, 280)
        # plt.ylim(100, 800)
        # plt.savefig("output/ori_x_y_small"+type)
        plt.close()
        # UR光标原始数据空间轨迹图
        plt.plot(raw_data["ori_x"], raw_data["ori_y"], color = colors[0])
        plt.plot(raw_data["ori_gaze_x"], raw_data["ori_gaze_y"], color = colors[1])
        plt.xlabel("x")
        plt.ylabel("y")

        plt.savefig(os.path.join(path_figsave,"ori_x_y_circle" + type))
        plt.close()

    """ 数据填充 """
    for i in range(raw_data_len):
        target_ball_point = torch.tensor(
            [raw_data["target_x"][i], raw_data["target_y"][i]]
        )
        target_ball_point_regualer = target_ball_point - center_point
        # 数据填充
        if raw_data["cursor_x"][i] is None:
            step = 0
            while raw_data["cursor_x"][i + step] is None and i + step + 1 < len(
                raw_data["cursor_x"]
            ):
                step += 1

            if i == 0:
                # padding from next
                for j in range(i, i + step):
                    raw_data["cursor_x"][j] = raw_data["cursor_x"][i + step]
                    raw_data["cursor_y"][j] = raw_data["cursor_y"][i + step]

            elif i + step == len(raw_data["cursor_x"]):
                # padding from pre
                for j in range(i, i + step):
                    raw_data["cursor_x"][j] = raw_data["cursor_x"][i - 1]
                    raw_data["cursor_y"][j] = raw_data["cursor_y"][i - 1]
            else:
                # average padding
                for j in range(i, i + step):
                    raw_data["cursor_x"][j] = raw_data["cursor_x"][i - 1] + (
                        (j - i + 1) / (step + 1)
                    ) * (raw_data["cursor_x"][i + step] - raw_data["cursor_x"][i - 1])
                    raw_data["cursor_y"][j] = raw_data["cursor_y"][i - 1] + (
                        (j - i + 1) / (step + 1)
                    ) * (raw_data["cursor_y"][i + step] - raw_data["cursor_y"][i - 1])

        track_ball_point = torch.tensor(
            [raw_data["cursor_x"][i], raw_data["cursor_y"][i]]
        )
        gaze_ball_point = torch.tensor(
            [raw_data["Gaze_x"][i], raw_data["Gaze_y"][i]]
        )
        track_ball_point_regualer = track_ball_point - center_point

        target_angel[i] = math.atan2(
            target_ball_point[1] - center_point[1],
            target_ball_point[0] - center_point[0],
        )

        ur_track_angel[i] = math.atan2(
            track_ball_point[1] - center_point[1],
            track_ball_point[0] - center_point[0],
        )
        gaze_track_angel[i] = math.atan2(
            gaze_ball_point[1] - center_point[1],
            gaze_ball_point[0] - center_point[0],
        )

        track_noise[i] = (
            distance_between_two_points((0, 0), track_ball_point_regualer)
            - radius_point
        ) # 径向 noise

    raw_data["pad_x"] = raw_data["cursor_x"]
    raw_data["pad_y"] = raw_data["cursor_y"]


    track_noise = 10.0 * F.tanh(0.01 * track_noise)
    # 一阶差分法计算target和x的切割点
    # 异或求差分交集
    # data_index_target=torch.zeros(raw_data_len)
    # data_index_x=torch.zeros(raw_data_len)
    diffs_target = torch.diff(target_angel)
    # plt.plot(diffs_target, label="x")
    # plt.show()
    # diffs_x=torch.diff(x_angel)

    start_point_target = torch.where(diffs_target < 0)[0] + 1
    # start_point_x = torch.where(diffs_x < 0)[0] + 1

    ''' 按周期分割数据片段 '''
    list_of_start_end = []
    list_of_target_angel = []
    list_of_urtrack_angel = []
    list_of_ur_error_angel = []
    list_of_ur_track_noise = []
    list_of_data = []
    list_of_ur_raw_x = []
    list_of_ur_raw_y = []

    list_of_gaze_track_angel = []
    list_of_gaze_error_angel = []
    # list_of_ur_track_noise = []
    # list_of_data = []
    list_of_gaze_raw_x = []
    list_of_gaze_raw_y = []
    first_start_point = -1
    last_end_point = -1
    for i in range(1, len(start_point_target) - 1): # 遍历每一个周期
        # data_index_target[start_point_target[i]:start_point_target[i+1]-1] = i
        current_start_point = (
            ur_track_angel[start_point_target[i] : start_point_target[i + 1] - 1].argmin()
            + start_point_target[i]
        ).item()
        if i == 1:
            first_start_point = current_start_point
        current_length = (
            ur_track_angel[current_start_point : start_point_target[i + 1] - 1].argmax()
            - 1
        )
        current_end_point = (current_start_point + current_length).item()
        if i == len(start_point_target) - 2:
            last_end_point = current_end_point

        current_target_split = target_angel[current_start_point:current_end_point]
        current_track_split = ur_track_angel[current_start_point:current_end_point]
        current_gaze_track_split = gaze_track_angel[current_start_point:current_end_point]

        list_of_start_end.append([current_start_point, current_end_point])

        list_of_target_angel.append(current_target_split)
        list_of_urtrack_angel.append(current_track_split)
        list_of_ur_error_angel.append(current_track_split - current_target_split)
        list_of_gaze_track_angel.append(current_gaze_track_split)
        list_of_gaze_error_angel.append(current_gaze_track_split - current_target_split)
        list_of_ur_track_noise.append(track_noise[current_start_point:current_end_point])
        list_of_ur_raw_x.append(
            torch.tensor(raw_data["cursor_x"][current_start_point:current_end_point])
        )
        list_of_ur_raw_y.append(
            torch.tensor(raw_data["cursor_y"][current_start_point:current_end_point])
        )
        list_of_gaze_raw_x.append(
            torch.tensor(raw_data["Gaze_x"][current_start_point:current_end_point])
        )
        list_of_gaze_raw_y.append(
            torch.tensor(raw_data["Gaze_y"][current_start_point:current_end_point])
        )

    ''' 低通滤波 '''
    # gaze track 数据
    entire_raw_gaze_x = (
        torch.tensor(raw_data["Gaze_x"][:-1], dtype=torch.float) - center_point[0]
    )
    entire_raw_gaze_x = util.low_pass_filter(entire_raw_gaze_x, 2, 120)


    entire_raw_gaze_y = (
        torch.tensor(raw_data["Gaze_y"][:-1], dtype=torch.float) - center_point[1]
    )

    entire_raw_gaze_y = util.low_pass_filter(entire_raw_gaze_y, 2, 120)

    entire_raw_gaze_vx = torch.diff(
        torch.tensor(
            raw_data["Gaze_x"],
            dtype=torch.float,
        )
    )

    entire_raw_gaze_vx = util.low_pass_filter(entire_raw_gaze_vx, 0.4, 120)

    entire_raw_gaze_vy = torch.diff(
        torch.tensor(
            raw_data["Gaze_y"],
            dtype=torch.float,
        )
    )

    entire_raw_gaze_vy = util.low_pass_filter(entire_raw_gaze_vy, 0.4, 120)

    raw_data_gx = (
        torch.tensor(raw_data["target_x"][:-1], dtype=torch.float) - center_point[0]
    )
    raw_data_gy = (
        torch.tensor(raw_data["target_y"][:-1], dtype=torch.float) - center_point[1]
    )
    # ur track 数据
    entire_raw_ur_x = (
        torch.tensor(raw_data["cursor_x"][:-1], dtype=torch.float) - center_point[0]
    )
    entire_raw_ur_x = util.low_pass_filter(entire_raw_ur_x, 0.5, 50)


    entire_raw_ur_y = (
        torch.tensor(raw_data["cursor_y"][:-1], dtype=torch.float) - center_point[1]
    )

    entire_raw_ur_y = util.low_pass_filter(entire_raw_ur_y, 0.5, 50)

    entire_raw_ur_vx = torch.diff(
        torch.tensor(
            raw_data["cursor_x"],
            dtype=torch.float,
        )
    )

    entire_raw_ur_vx = util.low_pass_filter(entire_raw_ur_vx, 0.4, 50)

    entire_raw_ur_vy = torch.diff(
        torch.tensor(
            raw_data["cursor_y"],
            dtype=torch.float,
        )
    )

    entire_raw_ur_vy = util.low_pass_filter(entire_raw_ur_vy, 0.4, 50)
    # 计算force范数
    # raw_data['force_norm'] = np.sqrt(raw_data['Hex_x']**2 + raw_data['Hex_y']**2)
    raw_data['force_norm'] = np.sqrt(np.array(raw_data['Hex_x'])**2 + np.array(raw_data['Hex_y'])**2)
    force_norm_tensor = torch.tensor(raw_data['force_norm'][:-1], dtype=torch.float) 
    force_norm_tensor = util.low_pass_filter(force_norm_tensor,0.1,20)

    ''' 计算 reacting time delay'''
    # calc gaze RT
    RT_r_gaze = torch.zeros(entire_raw_gaze_x.shape)
    angles_r = torch.atan2(entire_raw_gaze_y, entire_raw_gaze_x)
    angle_diffs_r = angles_r[1:] - angles_r[:-1] # 计算相邻元素差
    angle_diffs_r = (angle_diffs_r + np.pi) % (2 * np.pi) - np.pi # 归一到+-pi

    cumulative_angles_r = torch.cumsum(angle_diffs_r, dim=0)
    entire_raw_r = torch.cat((torch.tensor([0.0]), cumulative_angles_r))

    angles_gr = torch.atan2(raw_data_gy, raw_data_gx)
    angle_diffs_gr = angles_gr[1:] - angles_gr[:-1]
    angle_diffs_gr = (angle_diffs_gr + np.pi) % (2 * np.pi) - np.pi

    cumulative_angles_gr = torch.cumsum(angle_diffs_gr, dim=0)
    entire_raw_gr = torch.cat((torch.tensor([0.0]), cumulative_angles_gr))

    entire_raw_r[torch.where(entire_raw_r < entire_raw_gr.min())] = entire_raw_gr.min()
    entire_raw_r[torch.where(entire_raw_r > entire_raw_gr.max())] = entire_raw_gr.max()

    for i in range(len(entire_raw_gaze_x) - 1):
        curve = entire_raw_gr - entire_raw_r[i]
        diff = curve[1:] * curve[:-1]
        zero_crossings = torch.where(diff <= 0)[0]
        RT_r_gaze[i] = i - zero_crossings[torch.argmin(torch.abs(zero_crossings - i))]

    # calc ur RT
    RT_r_ur = torch.zeros(entire_raw_ur_x.shape)
    angles_r_ur = torch.atan2(entire_raw_ur_y, entire_raw_ur_x)
    angle_diffs_r_ur = angles_r_ur[1:] - angles_r_ur[:-1] # 计算相邻元素差
    angle_diffs_r_ur = (angle_diffs_r_ur + np.pi) % (2 * np.pi) - np.pi # 归一到+-pi

    cumulative_angles_r_ur = torch.cumsum(angle_diffs_r_ur, dim=0)
    entire_raw_r_ur = torch.cat((torch.tensor([0.0]), cumulative_angles_r_ur))

    entire_raw_r_ur[torch.where(entire_raw_r_ur < entire_raw_gr.min())] = entire_raw_gr.min()
    entire_raw_r_ur[torch.where(entire_raw_r_ur > entire_raw_gr.max())] = entire_raw_gr.max()

    for i in range(len(entire_raw_ur_x) - 1):
        curve_ur = entire_raw_gr - entire_raw_r_ur[i]
        diff = curve_ur[1:] * curve_ur[:-1]
        zero_crossings = torch.where(diff <= 0)[0]
        RT_r_ur[i] = i - zero_crossings[torch.argmin(torch.abs(zero_crossings - i))]

    ''' 计算 tracking error '''
    radius_xy = torch.sqrt(entire_raw_gaze_x * entire_raw_gaze_x + entire_raw_gaze_y * entire_raw_gaze_y)
    radius_gxgy = torch.sqrt(raw_data_gx * raw_data_gx + raw_data_gy * raw_data_gy)
    TE = radius_xy - radius_gxgy
    radius_xy_ur = torch.sqrt(entire_raw_ur_x * entire_raw_ur_x + entire_raw_ur_y * entire_raw_ur_y)
    
    TE_ur = radius_xy_ur - radius_gxgy

    ''' 计算 periodic increment(差分量) '''
    point_diff_x = entire_raw_gaze_x[1:] - entire_raw_gaze_x[:-1]
    point_diff_y = entire_raw_gaze_y[1:] - entire_raw_gaze_y[:-1]
    point_diff = torch.sqrt(point_diff_x * point_diff_x + point_diff_y * point_diff_y)
    #PI = torch.cat((torch.tensor([0.0]), point_diff))
    PI = torch.cat((point_diff[:1], point_diff))
    PI = util.low_pass_filter(PI, 0.1, 20)

    # raw_data["RT_r"] = RT_r.cpu().detach().numpy()
    # raw_data["TE"] = TE.cpu().detach().numpy()
    raw_data["filter_gaze_x"] = entire_raw_gaze_x.cpu().detach().numpy()
    raw_data["filter_gaze_y"] = entire_raw_gaze_y.cpu().detach().numpy()
    raw_data["filter_gaze_vx"] = entire_raw_gaze_vx.cpu().detach().numpy()
    raw_data["filter_gaze_vy"] = entire_raw_gaze_vy.cpu().detach().numpy()
    raw_data["filter_ur_x"] = entire_raw_ur_x.cpu().detach().numpy()
    raw_data["filter_ur_y"] = entire_raw_ur_y.cpu().detach().numpy()
    raw_data["filter_ur_vx"] = entire_raw_ur_vx.cpu().detach().numpy()
    raw_data["filter_ur_vy"] = entire_raw_ur_vy.cpu().detach().numpy()

    if Debug:
        plt.plot(raw_data["filter_gaze_x"], label="x")
        plt.xlabel("Frame")
        plt.plot(raw_data["filter_gaze_y"], label="y")
        plt.ylabel("Value")

        plt.savefig(os.path.join(path_figsave,"filter_gaze_x_y" + type))
        # plt.xlim(220, 280)
        # plt.ylim(100, 800)
        # plt.savefig("output/filter_x_y_small" + type)
        plt.close()

        plt.plot(raw_data["filter_ur_x"], label="x")
        plt.xlabel("Frame")
        plt.plot(raw_data["filter_ur_y"], label="y")
        plt.ylabel("Value")

        plt.savefig(os.path.join(path_figsave,"filter_ur_x_y" + type))
        # plt.xlim(220, 280)
        # plt.ylim(100, 800)
        # plt.savefig("output/filter_x_y_small" + type)
        plt.close()

        plt.plot(raw_data["filter_gaze_x"], raw_data["filter_gaze_y"])
        plt.plot(raw_data["filter_ur_x"], raw_data["filter_ur_y"])
        plt.xlabel("x") 
        plt.ylabel("y")

        plt.savefig(os.path.join(path_figsave,"filter_x_y_circle" + type))
        plt.close()

    drawfig = False
    if Debug:
        plt.subplot(4,1,1)
        # plt.plot(raw_data["stage_id"], label="stage_id")
        # plt.plot(raw_data["congnitive_flag"], label="cognitive_flag")
        # plt.plot(raw_data["motor_flag"], label="motor_flag")
        plt.xlabel("Frame")
        plt.ylabel("Value")
        plt.legend()
        plt.subplot(4,1,2)
        plt.plot(TE_ur, label="TEur")
        plt.plot(TE,label="TEgaze")
        plt.legend()
        plt.ylabel("Tracking Error")

        plt.subplot(4,1,3)
        plt.plot(raw_data["filter_ur_x"], label="x")
        plt.xlabel("Frame")
        plt.plot(raw_data["filter_ur_y"], label="y")
        plt.ylabel("Value")
        plt.ylim(-400,400)
        plt.subplot(4,1,4)
        plt.plot(raw_data["filter_gaze_x"], label="gazex")
        plt.xlabel("Frame")
        plt.plot(raw_data["filter_gaze_y"], label="gazey")
        plt.ylabel("value")
        plt.ylim(-400,400)
        plt.show()
        plt.savefig(os.path.join(path_figsave,"stage_fig" + type))
        plt.close()


    return {
        "raw_data": raw_data,
        "start_end": list_of_start_end,
        "target_angel": list_of_target_angel,
        "urtrack_angel": list_of_urtrack_angel,
        "error_angel": list_of_ur_error_angel,
        "track_noise": list_of_ur_track_noise,  
        "raw_x": list_of_ur_raw_x,
        "raw_y": list_of_ur_raw_y,
        "entire_raw_x": entire_raw_ur_x,
        "entire_raw_y": entire_raw_ur_y,
        "entire_raw_vx": entire_raw_ur_vx,
        "entire_raw_vy": entire_raw_ur_vy,
        "RT_ur": RT_r_ur,
        "TE_ur": TE_ur,
        "PI_ur": PI,
        "force_norm": force_norm_tensor,

        "gazetrack_angel": list_of_gaze_track_angel,
        "error_angel": list_of_gaze_error_angel,
        
        "raw_gaze_x": list_of_gaze_raw_x,
        "raw_gaze_y": list_of_gaze_raw_y,
        "entire_raw_gaze_x": entire_raw_gaze_x,
        "entire_raw_gaze_y": entire_raw_gaze_y,
        "entire_raw_gaze_vx": entire_raw_gaze_vx,
        "entire_raw_gaze_vy": entire_raw_gaze_vy,
        "RT_gaze": RT_r_gaze,
        "TE_gaze": TE,
        "PI_gaze": PI,
    }


if __name__ == "__main__":
    xk_N_AM = get_urtracker_data(
        r"./output/1739884039.4549184_normal_attention_motin_unglass.xlsx", type="Norm_AM", Debug=False
    )
    xk_ABN_AM = get_urtracker_data(
        r"./output/1739885879.2379827_distraction_attention_motion_unglass.xlsx", type="ABN_A", Debug=False
    )
    
    
    xk_N_gaze = torch.cat(
        [
            xk_N_AM["RT_gaze"].unsqueeze(0),
            xk_N_AM["TE_gaze"].unsqueeze(0),
            xk_N_AM["PI_gaze"].unsqueeze(0),
        ],
        dim=0,
    )
    print("xk_gaze shape:", xk_N_gaze.shape)
    
    xk_motion = torch.cat(
        [
            xk_N_AM["RT_ur"].unsqueeze(0),
            xk_N_AM["TE_ur"].unsqueeze(0),
            xk_N_AM["force_norm"].unsqueeze(0),
        ]
    )
    print("xk_motion shape:", xk_motion.shape)



    # 绘图debug
    draw_debug = True

    if draw_debug:
        # plt.close()
        show_length = 5000
        plt.figure("xk_N_gaze")
        plt.plot(xk_N_gaze[0, :].cpu(), label="RT")
        plt.plot(xk_N_gaze[1, :].cpu(), label="TE")
        plt.plot(xk_N_gaze[2, :].cpu(), label="PI")
        
        plt.legend()
        plt.show()

        fig = plt.figure("gaze")
        ax = fig.add_subplot(111, projection='3d')
        z = xk_N_gaze[2, :].cpu()
        x = xk_N_AM["raw_data"]['filter_gaze_x']
        y = xk_N_AM["raw_data"]['filter_gaze_y']
        ax.plot(x, y, z, label='parametric curve')
        ax.legend()
        plt.show()

        plt.figure("xk_N_motion")
        plt.plot(xk_motion[0, :show_length].cpu(), label="RT_ur")
        plt.plot(xk_motion[1, :show_length].cpu(), label="TE_ur")
        plt.plot(xk_motion[2, :show_length].cpu(), label="force")
        plt.legend()
        plt.show()

        fig = plt.figure("motion")
        ax = fig.add_subplot(111, projection='3d')
        z = xk_motion[2, :].cpu()
        x = xk_N_AM["raw_data"]['filter_ur_x']
        y = xk_N_AM["raw_data"]['filter_ur_y']
        ax.plot(x, y, z, label='parametric curve')
        ax.legend()
        plt.show()

        # plt.figure("....realtime")
        # plt.plot(xk_N_gaze[2, :show_length].cpu(), label="gaze")
        # plt.plot(xk_motion[2, :show_length].cpu(), label="motion")
        # plt.legend()
        # plt.show()
import torch
import torch.nn.functional as F
from openpyxl import load_workbook
import math
import matplotlib.pyplot as plt
import util
import copy
import numpy as np

# 是用来描述动态系统的，用来在后边去模拟正常状态、故障状态下系统每个周期的输出
g = 9.8
Sp = 5e-5
As = 1 / 0.0154
Cs = torch.tensor([1, 0.8, 1], dtype=torch.float64)
alpha = torch.tensor([0.8, 0.9], dtype=torch.float64)  # Fault 1&2
Sf = torch.tensor([1e-6, 2e-6], dtype=torch.float64)  # fault 3&4
w1, w2 = 0.3, 0.3  # u1&u2，是用来生成数据的参数


def get_normal_mode(x0, n, steps, start_k=0):
    # 初始化记录水箱水位的数组
    x = x0.clone()
    ut = torch.tensor(1.2e-4)
    u1 = torch.tensor(0.0)
    u2 = torch.tensor(0.0)
    xk0 = torch.zeros((n, steps), dtype=torch.float64)
    # 循环遍历每个时间步
    for k in range(start_k, steps):
        # k-1
        q13 = (
                Cs[0]
                * Sp
                * torch.sign(x[0] - x[2])
                * torch.sqrt(2 * g * torch.abs(x[0] - x[2]))
        )
        q32 = (
                Cs[2]
                * Sp
                * torch.sign(x[2] - x[1])
                * torch.sqrt(2 * g * torch.abs(x[2] - x[1]))
        )
        q20 = Cs[1] * Sp * torch.sqrt(2 * g * x[1])
        u1 = ((torch.sign(u1) * torch.sign(ut - u1) + 1) / 2) * (
                -5 * Sp * (x[0] - 0.5)
                + 0.8 * Sp * (1.5 + torch.sin(torch.tensor(w1 * (k + 1))))
        )
        u2 = ((torch.sign(u2) * torch.sign(ut - u2) + 1) / 2) * (
                -5 * Sp * (x[1] - 0.5)
                + 0.8 * Sp * (1.5 + torch.cos(torch.tensor(w2 * (k + 1))))
        )
        # k
        x[0] += As * (-q13 + u1)
        x[1] += As * (q32 - q20 + u2)
        x[2] += As * (q13 - q32)
        xk0[:, k] = x.view(-1)
    return xk0


def get_fault_mode_1(x0, n, steps, start_k=0):
    x = x0.clone()
    ut = torch.tensor(1.2e-4)
    u1 = torch.tensor(0.0)
    u2 = torch.tensor(0.0)
    xk0 = torch.zeros((n, steps), dtype=torch.float64)

    for k in range(start_k, steps):
        q13 = (
                Cs[0]
                * Sp
                * torch.sign(x[0] - x[2])
                * torch.sqrt(2 * g * torch.abs(x[0] - x[2]))
        )
        q32 = (
                Cs[2]
                * Sp
                * torch.sign(x[2] - x[1])
                * torch.sqrt(2 * g * torch.abs(x[2] - x[1]))
        )
        q20 = Cs[1] * Sp * torch.sqrt(2 * g * x[1])

        u1 = ((torch.sign(u1) * torch.sign(ut - u1) + 1) / 2) * (
                -5 * Sp * (x[0] - 0.5)
                + 0.8 * Sp * (1.5 + torch.sin(torch.tensor(w1 * (k + 1))))
        )
        u1 *= alpha[0]

        u2 = ((torch.sign(u2) * torch.sign(ut - u2) + 1) / 2) * (
                -5 * Sp * (x[1] - 0.5)
                + 0.8 * Sp * (1.5 + torch.cos(torch.tensor(w2 * (k + 1))))
        )

        x[0] += As * (-q13 + u1)
        x[1] += As * (q32 - q20 + u2)
        x[2] += As * (q13 - q32)
        xk0[:, k] = x.view(-1)
    return xk0


def get_fault_mode_2(x0, n, steps, start_k=0):
    x = x0.clone()
    ut = torch.tensor(1.2e-4)
    u1 = torch.tensor(0.0)
    u2 = torch.tensor(0.0)
    xk0 = torch.zeros((n, steps), dtype=torch.float64)

    for k in range(start_k, steps):
        q13 = (
                Cs[0]
                * Sp
                * torch.sign(x[0] - x[2])
                * torch.sqrt(2 * g * torch.abs(x[0] - x[2]))
        )
        q32 = (
                Cs[2]
                * Sp
                * torch.sign(x[2] - x[1])
                * torch.sqrt(2 * g * torch.abs(x[2] - x[1]))
        )
        q20 = Cs[1] * Sp * torch.sqrt(2 * g * x[1])

        u1 = ((torch.sign(u1) * torch.sign(ut - u1) + 1) / 2) * (
                -5 * Sp * (x[0] - 0.5)
                + 0.8 * Sp * (1.5 + torch.sin(torch.tensor(w1 * (k + 1))))
        )

        u2 = ((torch.sign(u2) * torch.sign(ut - u2) + 1) / 2) * (
                -5 * Sp * (x[1] - 0.5)
                + 0.8 * Sp * (1.5 + torch.cos(torch.tensor(w2 * (k + 1))))
        )
        u2 *= alpha[1]

        x[0] += As * (-q13 + u1)
        x[1] += As * (q32 - q20 + u2)
        x[2] += As * (q13 - q32)
        xk0[:, k] = x.view(-1)
    return xk0


import torch

# Assuming the global variables are defined in the scope where these functions are called
g = 9.8
Sp = 5e-5
As = 1 / 0.0154
Cs = torch.tensor([1, 0.8, 1], dtype=torch.float64)
Sf = torch.tensor([1e-6, 2e-6], dtype=torch.float64)  # fault 3&4
w1, w2 = 0.3, 0.3  # u1&u2


def get_fault_mode_3(x0, n, steps, start_k=0):
    x = x0.clone()
    ut = torch.tensor(1.2e-4)
    u1 = torch.tensor(0.0)
    u2 = torch.tensor(0.0)
    xk0 = torch.zeros((n, steps), dtype=torch.float64)

    for k in range(start_k, steps):
        q13 = (
                Cs[0]
                * Sp
                * torch.sign(x[0] - x[2])
                * torch.sqrt(2 * g * torch.abs(x[0] - x[2]))
        )
        q32 = (
                Cs[2]
                * Sp
                * torch.sign(x[2] - x[1])
                * torch.sqrt(2 * g * torch.abs(x[2] - x[1]))
        )
        q20 = Cs[1] * Sp * torch.sqrt(2 * g * x[1])
        q1f = Cs[0] * Sf[0] * torch.sqrt(2 * g * x[0])

        u1 = ((torch.sign(u1) * torch.sign(ut - u1) + 1) / 2) * (
                -5 * Sp * (x[0] - 0.5)
                + 0.8 * Sp * (1.5 + torch.sin(torch.tensor(w1 * (k + 1))))
        )
        u2 = ((torch.sign(u2) * torch.sign(ut - u2) + 1) / 2) * (
                -5 * Sp * (x[1] - 0.5)
                + 0.8 * Sp * (1.5 + torch.cos(torch.tensor(w2 * (k + 1))))
        )

        x[0] += As * (-q13 - q1f + u1)
        x[1] += As * (q32 - q20 + u2)
        x[2] += As * (q13 - q32)
        xk0[:, k] = x.view(-1)
    return xk0


def get_fault_mode_4(x0, n, steps, start_k=0):
    x = x0.clone()
    ut = torch.tensor(1.2e-4)
    u1 = torch.tensor(0.0)
    u2 = torch.tensor(0.0)
    xk0 = torch.zeros((n, steps), dtype=torch.float64)

    for k in range(start_k, steps):
        q13 = (
                Cs[0]
                * Sp
                * torch.sign(x[0] - x[2])
                * torch.sqrt(2 * g * torch.abs(x[0] - x[2]))
        )
        q32 = (
                Cs[2]
                * Sp
                * torch.sign(x[2] - x[1])
                * torch.sqrt(2 * g * torch.abs(x[2] - x[1]))
        )
        q20 = Cs[1] * Sp * torch.sqrt(2 * g * x[1])
        q2f = Cs[1] * Sf[1] * torch.sqrt(2 * g * x[1])

        u1 = ((torch.sign(u1) * torch.sign(ut - u1) + 1) / 2) * (
                -5 * Sp * (x[0] - 0.5)
                + 0.8 * Sp * (1.5 + torch.sin(torch.tensor(w1 * (k + 1))))
        )
        u2 = ((torch.sign(u2) * torch.sign(ut - u2) + 1) / 2) * (
                -5 * Sp * (x[1] - 0.5)
                + 0.8 * Sp * (1.5 + torch.cos(torch.tensor(w2 * (k + 1))))
        )

        x[0] += As * (-q13 + u1)
        x[1] += As * (q32 - q20 - q2f + u2)
        x[2] += As * (q13 - q32)
        xk0[:, k] = x.view(-1)
    return xk0


def get_test_data(x0, n, steps, fault_type, fault_time):
    x = torch.zeros([n, steps], dtype=torch.float64)
    x_normal = get_normal_mode(x0, n, steps)
    x[:, :fault_time] = x_normal[:, :fault_time]
    if fault_type == 0:
        return x_normal, x_normal
    elif fault_type == 1:
        fault_data = get_fault_mode_1(x[:, fault_time].clone(), n, steps, fault_time)
        x[:, fault_time:] = fault_data[:, fault_time:]
        return x, x_normal
    elif fault_type == 2:
        fault_data = get_fault_mode_2(x[:, fault_time].clone(), n, steps, fault_time)
        x[:, fault_time:] = fault_data[:, fault_time:]
        return x, x_normal
    elif fault_type == 3:
        fault_data = get_fault_mode_3(x[:, fault_time].clone(), n, steps, fault_time)
        x[:, fault_time:] = fault_data[:, fault_time:]
        return x, x_normal
    elif fault_type == 4:
        fault_data = get_fault_mode_4(x[:, fault_time].clone(), n, steps, fault_time)
        x[:, fault_time:] = fault_data[:, fault_time:]
        return x, x_normal
    else:
        print("unknown error code, please verify the data generator")


def distance_between_two_points(pointA, pointB):
    x1, y1 = pointA
    x2, y2 = pointB
    return math.sqrt((x2 - x1) ** 2 + (y2 - y1) ** 2)


def get_trackball_data(path="output/trackballdata.xlsx"):
    # x = torch.zeros([n, steps], dtype=torch.float64)
    # tobii on surface
    center_point = torch.tensor([823, 548])
    # center_point = torch.tensor([300, 300])
    radius_point = 379

    workbook = load_workbook(filename=path)

    sheet = workbook.active

    # 读取数据
    raw_data = {}
    keys = []  # ['time', 'target_x', 'target_y', 'cursor_x', 'cursor_y']

    for row in sheet.iter_rows(values_only=True):
        # data.append(row)
        if len(raw_data.keys()) == 0:
            for key in row:
                if key is not None:
                    keys.append(key)
                    raw_data[key] = []
            continue

        for i in range(len(keys)):
            raw_data[keys[i]].append(row[i + 1])

    raw_data_ori = copy.deepcopy(raw_data)
    raw_data_len = len(raw_data["time"])
    target_angel = torch.zeros(raw_data_len)
    track_angel = torch.zeros(raw_data_len)
    track_noise = torch.zeros(raw_data_len)

    raw_x = torch.zeros(raw_data_len)
    raw_y = torch.zeros(raw_data_len)

    raw_data["ori_x"] = raw_data["cursor_x"]
    raw_data["ori_y"] = raw_data["cursor_y"]

    plt.plot(raw_data["ori_x"], label="x")
    plt.xlabel("Frame")
    plt.plot(raw_data["ori_y"], label="y")
    plt.ylabel("Value")

    plt.savefig("ori_x_y")
    plt.xlim(220, 280)
    plt.ylim(100, 800)
    plt.savefig("ori_x_y_small")
    plt.close()

    plt.plot(raw_data["ori_x"], raw_data["ori_y"])
    plt.xlabel("x")
    plt.ylabel("y")

    plt.savefig("ori_x_y_circle")
    plt.close()

    for i in range(raw_data_len):
        target_ball_point = torch.tensor(
            [raw_data["target_x"][i], raw_data["target_y"][i]]
        )
        target_ball_point_regualer = target_ball_point - center_point
        if raw_data["cursor_x"][i] is None:
            step = 0
            while raw_data["cursor_x"][i + step] is None and i + step + 1 < len(
                    raw_data["cursor_x"]
            ):
                step += 1

            if i == 0:
                # padding from next
                for j in range(i, i + step):
                    raw_data["cursor_x"][j] = raw_data["cursor_x"][i + step]
                    raw_data["cursor_y"][j] = raw_data["cursor_y"][i + step]

            elif i + step == len(raw_data["cursor_x"]):
                # padding from pre
                for j in range(i, i + step):
                    raw_data["cursor_x"][j] = raw_data["cursor_x"][i - 1]
                    raw_data["cursor_y"][j] = raw_data["cursor_y"][i - 1]
            else:
                # average padding
                for j in range(i, i + step):
                    raw_data["cursor_x"][j] = raw_data["cursor_x"][i - 1] + (
                            (j - i + 1) / (step + 1)
                    ) * (
                                                         raw_data["cursor_x"][i + step]
                                                         - raw_data["cursor_x"][i - 1]
                                                 )
                    raw_data["cursor_y"][j] = raw_data["cursor_y"][i - 1] + (
                            (j - i + 1) / (step + 1)
                    ) * (
                                                         raw_data["cursor_y"][i + step]
                                                         - raw_data["cursor_y"][i - 1]
                                                 )

        track_ball_point = torch.tensor(
            [raw_data["cursor_x"][i], raw_data["cursor_y"][i]]
        )
        track_ball_point_regualer = track_ball_point - center_point

        target_angel[i] = math.atan2(
            target_ball_point[1] - center_point[1],
            target_ball_point[0] - center_point[0],
        )

        track_angel[i] = math.atan2(
            track_ball_point[1] - center_point[1],
            track_ball_point[0] - center_point[0],
        )

        track_noise[i] = (
                distance_between_two_points((0, 0), track_ball_point_regualer)
                - radius_point
        )

    raw_data["pad_x"] = raw_data["cursor_x"]
    raw_data["pad_y"] = raw_data["cursor_y"]

    plt.plot(raw_data["pad_x"], label="x")
    plt.xlabel("Frame")
    plt.plot(raw_data["pad_y"], label="y")
    plt.ylabel("Value")

    plt.savefig("pad_x_y")
    plt.xlim(220, 280)
    plt.ylim(100, 800)
    plt.savefig("pad_x_y_small")
    plt.close()

    plt.plot(raw_data["pad_x"], raw_data["pad_y"])
    plt.xlabel("x")
    plt.ylabel("y")

    plt.savefig("pad_x_y_circle")
    plt.close()

    track_noise = 10.0 * F.tanh(0.01 * track_noise)
    # 一阶差分法计算target和x的切割点
    # 异或求差分交集
    # data_index_target=torch.zeros(raw_data_len)
    # data_index_x=torch.zeros(raw_data_len)
    diffs_target = torch.diff(target_angel)
    # diffs_x=torch.diff(x_angel)

    start_point_target = torch.where(diffs_target < 0)[0] + 1
    # start_point_x = torch.where(diffs_x < 0)[0] + 1

    list_of_start_end = []
    list_of_target_angel = []
    list_of_track_angel = []
    list_of_error_angel = []
    list_of_track_noise = []
    list_of_data = []
    list_of_raw_x = []
    list_of_raw_y = []
    first_start_point = -1
    last_end_point = -1
    for i in range(1, len(start_point_target) - 1):
        # data_index_target[start_point_target[i]:start_point_target[i+1]-1] = i
        current_start_point = (
                track_angel[start_point_target[i]: start_point_target[i + 1] - 1].argmin()
                + start_point_target[i]
        ).item()
        if i == 1:
            first_start_point = current_start_point
        current_length = (
                track_angel[current_start_point: start_point_target[i + 1] - 1].argmax()
                - 1
        )
        current_end_point = (current_start_point + current_length).item()
        if i == len(start_point_target) - 2:
            last_end_point = current_end_point

        current_target_split = target_angel[current_start_point:current_end_point]
        current_track_split = track_angel[current_start_point:current_end_point]

        list_of_start_end.append([current_start_point, current_end_point])

        list_of_target_angel.append(current_target_split)
        list_of_track_angel.append(current_track_split)
        list_of_error_angel.append(current_track_split - current_target_split)
        list_of_track_noise.append(track_noise[current_start_point:current_end_point])
        list_of_raw_x.append(
            torch.tensor(raw_data["cursor_x"][current_start_point:current_end_point])
        )
        list_of_raw_y.append(
            torch.tensor(raw_data["cursor_y"][current_start_point:current_end_point])
        )

    entire_raw_x = (
            torch.tensor(
                raw_data["cursor_x"][:-1], dtype=torch.float
            )
            - center_point[0]
    )

    entire_raw_x = util.low_pass_filter(entire_raw_x, 0.5, 50)

    entire_raw_y = (
            torch.tensor(
                raw_data["cursor_y"][:-1], dtype=torch.float
            )
            - center_point[1]
    )

    entire_raw_y = util.low_pass_filter(entire_raw_y, 0.5, 50)

    entire_raw_vx = torch.diff(
        torch.tensor(
            raw_data["cursor_x"],
            dtype=torch.float,
        )
    )

    entire_raw_vx = util.low_pass_filter(entire_raw_vx, 0.4, 50)

    entire_raw_vy = torch.diff(
        torch.tensor(
            raw_data["cursor_y"],
            dtype=torch.float,
        )
    )

    entire_raw_vy = util.low_pass_filter(entire_raw_vy, 0.4, 50)

    raw_data["filter_x"] = entire_raw_x.cpu().detach().numpy()
    raw_data["filter_y"] = entire_raw_y.cpu().detach().numpy()
    raw_data["filter_vx"] = entire_raw_vx.cpu().detach().numpy()
    raw_data["filter_vy"] = entire_raw_vy.cpu().detach().numpy()

    plt.plot(raw_data["filter_x"], label="x")
    plt.xlabel("Frame")
    plt.plot(raw_data["filter_y"], label="y")
    plt.ylabel("Value")

    plt.savefig("output/filter_x_y")
    plt.xlim(220, 280)
    plt.ylim(100, 800)
    plt.savefig("output/filter_x_y_small")
    plt.close()

    plt.plot(raw_data["filter_x"], raw_data["filter_y"])
    plt.xlabel("x")
    plt.ylabel("y")

    plt.savefig("output/filter_x_y_circle")
    plt.close()

    return {
        "raw_data": raw_data,
        "start_end": list_of_start_end,
        "target_angel": list_of_target_angel,
        "track_angel": list_of_track_angel,
        "error_angel": list_of_error_angel,
        "track_noise": list_of_track_noise,
        "raw_x": list_of_raw_x,
        "raw_y": list_of_raw_y,
        "entire_raw_x": entire_raw_x,
        "entire_raw_y": entire_raw_y,
        "entire_raw_vx": entire_raw_vx,
        "entire_raw_vy": entire_raw_vy,
    }


def get_urtracker_data_test(path="output/trackballdata.xlsx", type = "mode1"):
    # x = torch.zeros([n, steps], dtype=torch.float64)
    # tobii on surface
    center_point = torch.tensor([432, 432])
    # center_point = torch.tensor([300, 300])
    radius_point = 252

    workbook = load_workbook(filename=path)

    sheet = workbook.active

    # 读取数据
    raw_data = {}
    keys = []  # ['time', 'target_x', 'target_y', 'cursor_x', 'cursor_y']
    '''
    ['time', 'target_x', 'target_y', 'cursor_x', 'cursor_y', 'Hex_x', 'Hex_y', 'Hex_z', 
    'linear_x', 'linear_y', 'linear_z', 'pose_x', 'pose_y', 'pose_z', 'pose_rx', 'pose_ry', 'pose_rz']
    '''

    for row in sheet.iter_rows(values_only=True):
        # data.append(row)
        if len(raw_data.keys()) == 0:
            for key in row:
                if key is not None:
                    keys.append(key)
                    raw_data[key] = []
            continue

        for i in range(len(keys)):
            raw_data[keys[i]].append(row[i + 1])

    raw_data_ori = copy.deepcopy(raw_data)
    raw_data_len = len(raw_data["time"])
    target_angel = torch.zeros(raw_data_len)
    track_angel = torch.zeros(raw_data_len)
    track_noise = torch.zeros(raw_data_len)

    raw_x = torch.zeros(raw_data_len)
    raw_y = torch.zeros(raw_data_len)

    raw_data["ori_x"] = raw_data["cursor_x"]
    raw_data["ori_y"] = raw_data["cursor_y"]

    plt.plot(raw_data["ori_x"], label="x")
    plt.xlabel("Frame")
    plt.plot(raw_data["ori_y"], label="y")
    plt.ylabel("Value")

    plt.savefig("output/ori_x_y"+type)
    plt.xlim(220, 280)
    plt.ylim(100, 800)
    plt.savefig("output/ori_x_y_small"+type)
    plt.close()

    plt.plot(raw_data["ori_x"], raw_data["ori_y"])
    plt.xlabel("x")
    plt.ylabel("y")

    plt.savefig("output/ori_x_y_circle"+type)
    plt.close()

    for i in range(raw_data_len):
        target_ball_point = torch.tensor(
            [raw_data["target_x"][i], raw_data["target_y"][i]]
        )
        target_ball_point_regualer = target_ball_point - center_point
        if raw_data["cursor_x"][i] is None:
            step = 0
            while raw_data["cursor_x"][i + step] is None and i + step + 1 < len(
                    raw_data["cursor_x"]
            ):
                step += 1

            if i == 0:
                # padding from next
                for j in range(i, i + step):
                    raw_data["cursor_x"][j] = raw_data["cursor_x"][i + step]
                    raw_data["cursor_y"][j] = raw_data["cursor_y"][i + step]

            elif i + step == len(raw_data["cursor_x"]):
                # padding from pre
                for j in range(i, i + step):
                    raw_data["cursor_x"][j] = raw_data["cursor_x"][i - 1]
                    raw_data["cursor_y"][j] = raw_data["cursor_y"][i - 1]
            else:
                # average padding
                for j in range(i, i + step):
                    raw_data["cursor_x"][j] = raw_data["cursor_x"][i - 1] + (
                            (j - i + 1) / (step + 1)
                    ) * (
                                                         raw_data["cursor_x"][i + step]
                                                         - raw_data["cursor_x"][i - 1]
                                                 )
                    raw_data["cursor_y"][j] = raw_data["cursor_y"][i - 1] + (
                            (j - i + 1) / (step + 1)
                    ) * (
                                                         raw_data["cursor_y"][i + step]
                                                         - raw_data["cursor_y"][i - 1]
                                                 )

        track_ball_point = torch.tensor(
            [raw_data["cursor_x"][i], raw_data["cursor_y"][i]]
        )
        track_ball_point_regualer = track_ball_point - center_point

        target_angel[i] = math.atan2(
            target_ball_point[1] - center_point[1],
            target_ball_point[0] - center_point[0],
        )

        track_angel[i] = math.atan2(
            track_ball_point[1] - center_point[1],
            track_ball_point[0] - center_point[0],
        )

        track_noise[i] = (
                distance_between_two_points((0, 0), track_ball_point_regualer)
                - radius_point
        )

    raw_data["pad_x"] = raw_data["cursor_x"]
    raw_data["pad_y"] = raw_data["cursor_y"]

    plt.plot(raw_data["pad_x"], label="x")
    plt.xlabel("Frame")
    plt.plot(raw_data["pad_y"], label="y")
    plt.ylabel("Value")

    plt.savefig("output/pad_x_y"+type)
    plt.xlim(220, 280)
    plt.ylim(100, 800)
    plt.savefig("output/pad_x_y_small"+type)
    plt.close()

    plt.plot(raw_data["pad_x"], raw_data["pad_y"])
    plt.xlabel("x")
    plt.ylabel("y")

    plt.savefig("output/pad_x_y_circle"+type)
    plt.close()

    track_noise = 10.0 * F.tanh(0.01 * track_noise)
    # 一阶差分法计算target和x的切割点
    # 异或求差分交集
    # data_index_target=torch.zeros(raw_data_len)
    # data_index_x=torch.zeros(raw_data_len)
    diffs_target = torch.diff(target_angel)
    # diffs_x=torch.diff(x_angel)

    start_point_target = torch.where(diffs_target < 0)[0] + 1
    # start_point_x = torch.where(diffs_x < 0)[0] + 1

    list_of_start_end = []
    list_of_target_angel = []
    list_of_track_angel = []
    list_of_error_angel = []
    list_of_track_noise = []
    list_of_data = []
    list_of_raw_x = []
    list_of_raw_y = []
    first_start_point = -1
    last_end_point = -1
    for i in range(1, len(start_point_target) - 1):
        # data_index_target[start_point_target[i]:start_point_target[i+1]-1] = i
        current_start_point = (
                track_angel[start_point_target[i]: start_point_target[i + 1] - 1].argmin()
                + start_point_target[i]
        ).item()
        if i == 1:
            first_start_point = current_start_point
        current_length = (
                track_angel[current_start_point: start_point_target[i + 1] - 1].argmax()
                - 1
        )
        current_end_point = (current_start_point + current_length).item()
        if i == len(start_point_target) - 2:
            last_end_point = current_end_point

        current_target_split = target_angel[current_start_point:current_end_point]
        current_track_split = track_angel[current_start_point:current_end_point]

        list_of_start_end.append([current_start_point, current_end_point])

        list_of_target_angel.append(current_target_split)
        list_of_track_angel.append(current_track_split)
        list_of_error_angel.append(current_track_split - current_target_split)
        list_of_track_noise.append(track_noise[current_start_point:current_end_point])
        list_of_raw_x.append(
            torch.tensor(raw_data["cursor_x"][current_start_point:current_end_point])
        )
        list_of_raw_y.append(
            torch.tensor(raw_data["cursor_y"][current_start_point:current_end_point])
        )

    entire_raw_x = (
            torch.tensor(
                raw_data["cursor_x"][:-1], dtype=torch.float
            )
            - center_point[0]
    )

    entire_raw_x = util.low_pass_filter(entire_raw_x, 0.5, 50)

    entire_raw_y = (
            torch.tensor(
                raw_data["cursor_y"][:-1], dtype=torch.float
            )
            - center_point[1]
    )

    entire_raw_y = util.low_pass_filter(entire_raw_y, 0.5, 50)

    entire_raw_vx = torch.diff(
        torch.tensor(
            raw_data["cursor_x"],
            dtype=torch.float,
        )
    )

    entire_raw_vx = util.low_pass_filter(entire_raw_vx, 0.4, 50)

    entire_raw_vy = torch.diff(
        torch.tensor(
            raw_data["cursor_y"],
            dtype=torch.float,
        )
    )

    entire_raw_vy = util.low_pass_filter(entire_raw_vy, 0.4, 50)

    raw_data_gx = torch.tensor(raw_data["target_x"][:-1], dtype=torch.float) - center_point[0]
    raw_data_gy = torch.tensor(raw_data["target_y"][:-1], dtype=torch.float) - center_point[1]

    # calc RT

    RT_r = torch.zeros(entire_raw_x.shape)

    angles_r = torch.atan2(entire_raw_y, entire_raw_x)
    angle_diffs_r = angles_r[1:] - angles_r[:-1]
    angle_diffs_r = (angle_diffs_r + np.pi) % (2 * np.pi) - np.pi

    cumulative_angles_r = torch.cumsum(angle_diffs_r, dim=0)
    entire_raw_r = torch.cat((torch.tensor([0.0]), cumulative_angles_r))

    angles_gr = torch.atan2(raw_data_gy, raw_data_gx)
    angle_diffs_gr = angles_gr[1:] - angles_gr[:-1]
    angle_diffs_gr = (angle_diffs_gr + np.pi) % (2 * np.pi) - np.pi

    cumulative_angles_gr = torch.cumsum(angle_diffs_gr, dim=0)
    entire_raw_gr = torch.cat((torch.tensor([0.0]), cumulative_angles_gr))

    entire_raw_r[torch.where(entire_raw_r < entire_raw_gr.min())] = entire_raw_gr.min()
    entire_raw_r[torch.where(entire_raw_r > entire_raw_gr.max())] = entire_raw_gr.max()

    for i in range(len(entire_raw_x) - 1):
        curve = entire_raw_gr - entire_raw_r[i]
        diff = curve[1:] * curve[:-1]
        zero_crossings = torch.where(diff <= 0)[0]
        RT_r[i] = i - zero_crossings[torch.argmin(torch.abs(zero_crossings - i))]

    # calc TE
    radius_xy = torch.sqrt(entire_raw_x * entire_raw_x + entire_raw_y * entire_raw_y)
    radius_gxgy = torch.sqrt(raw_data_gx * raw_data_gx + raw_data_gy * raw_data_gy)
    TE = radius_xy - radius_gxgy

    # calc PI
    point_diff_x = entire_raw_x[1:] - entire_raw_x[:-1]
    point_diff_y = entire_raw_y[1:] - entire_raw_y[:-1]
    point_diff = torch.sqrt(point_diff_x * point_diff_x + point_diff_y * point_diff_y)
    PI = torch.cat((torch.tensor([0.0]), point_diff))

    # raw_data["RT_r"] = RT_r.cpu().detach().numpy()
    # raw_data["TE"] = TE.cpu().detach().numpy()
    raw_data["filter_x"] = entire_raw_x.cpu().detach().numpy()
    raw_data["filter_y"] = entire_raw_y.cpu().detach().numpy()
    raw_data["filter_vx"] = entire_raw_vx.cpu().detach().numpy()
    raw_data["filter_vy"] = entire_raw_vy.cpu().detach().numpy()

    plt.plot(raw_data["filter_x"], label="x")
    plt.xlabel("Frame")
    plt.plot(raw_data["filter_y"], label="y")
    plt.ylabel("Value")

    plt.savefig("output/filter_x_y"+type)
    plt.xlim(220, 280)
    plt.ylim(100, 800)
    plt.savefig("output/filter_x_y_small"+type)
    plt.close()

    plt.plot(raw_data["filter_x"], raw_data["filter_y"])
    plt.xlabel("x")
    plt.ylabel("y")

    plt.savefig("output/filter_x_y_circle"+type)
    plt.close()

    return {
        "raw_data": raw_data,
        "start_end": list_of_start_end,
        "target_angel": list_of_target_angel,
        "track_angel": list_of_track_angel,
        "error_angel": list_of_error_angel,
        "track_noise": list_of_track_noise,
        "raw_x": list_of_raw_x,
        "raw_y": list_of_raw_y,
        "entire_raw_x": entire_raw_x,
        "entire_raw_y": entire_raw_y,
        "entire_raw_vx": entire_raw_vx,
        "entire_raw_vy": entire_raw_vy,
        "RT_r": RT_r,
        "TE": TE,
        "PI": PI
    }


def get_trackball_data_robio(path="output/trackballdata.xlsx", type = "mode1"):
    # x = torch.zeros([n, steps], dtype=torch.float64)
    # tobii on surface
    center_point = torch.tensor([432, 432])
    # center_point = torch.tensor([300, 300])
    radius_point = 252

    workbook = load_workbook(filename=path)

    sheet = workbook.active

    # 读取数据
    raw_data = {}
    keys = []  # ['time', 'target_x', 'target_y', 'cursor_x', 'cursor_y']
    '''
    [time, target_x, target_y, cursor_x, cursor_y, Hex_x, Hex_y, Hex_z, 
    linear_x, linear_y, linear_z, pose_x, pose_y, pose_z, pose_rx, pose_ry, pose_rz]
    '''

    for row in sheet.iter_rows(values_only=True):
        # data.append(row)
        if len(raw_data.keys()) == 0:
            for key in row:
                if key is not None:
                    keys.append(key)
                    raw_data[key] = []
            continue

        for i in range(len(keys)):
            raw_data[keys[i]].append(row[i + 1])

    raw_data_ori = copy.deepcopy(raw_data)
    raw_data_len = len(raw_data["time"])
    target_angel = torch.zeros(raw_data_len)
    track_angel = torch.zeros(raw_data_len)
    track_noise = torch.zeros(raw_data_len)

    raw_x = torch.zeros(raw_data_len)
    raw_y = torch.zeros(raw_data_len)

    raw_data["ori_x"] = raw_data["cursor_x"]
    raw_data["ori_y"] = raw_data["cursor_y"]

    plt.plot(raw_data["ori_x"], label="x")
    plt.xlabel("Frame")
    plt.plot(raw_data["ori_y"], label="y")
    plt.ylabel("Value")

    plt.savefig("output/ori_x_y"+type)
    plt.xlim(220, 280)
    plt.ylim(100, 800)
    plt.savefig("output/ori_x_y_small"+type)
    plt.close()

    plt.plot(raw_data["ori_x"], raw_data["ori_y"])
    plt.xlabel("x")
    plt.ylabel("y")

    plt.savefig("output/ori_x_y_circle"+type)
    plt.close()

    for i in range(raw_data_len):
        target_ball_point = torch.tensor(
            [raw_data["target_x"][i], raw_data["target_y"][i]]
        )
        target_ball_point_regualer = target_ball_point - center_point
        if raw_data["cursor_x"][i] is None:
            step = 0
            while raw_data["cursor_x"][i + step] is None and i + step + 1 < len(
                    raw_data["cursor_x"]
            ):
                step += 1

            if i == 0:
                # padding from next
                for j in range(i, i + step):
                    raw_data["cursor_x"][j] = raw_data["cursor_x"][i + step]
                    raw_data["cursor_y"][j] = raw_data["cursor_y"][i + step]

            elif i + step == len(raw_data["cursor_x"]):
                # padding from pre
                for j in range(i, i + step):
                    raw_data["cursor_x"][j] = raw_data["cursor_x"][i - 1]
                    raw_data["cursor_y"][j] = raw_data["cursor_y"][i - 1]
            else:
                # average padding
                for j in range(i, i + step):
                    raw_data["cursor_x"][j] = raw_data["cursor_x"][i - 1] + (
                            (j - i + 1) / (step + 1)
                    ) * (
                                                         raw_data["cursor_x"][i + step]
                                                         - raw_data["cursor_x"][i - 1]
                                                 )
                    raw_data["cursor_y"][j] = raw_data["cursor_y"][i - 1] + (
                            (j - i + 1) / (step + 1)
                    ) * (
                                                         raw_data["cursor_y"][i + step]
                                                         - raw_data["cursor_y"][i - 1]
                                                 )

        track_ball_point = torch.tensor(
            [raw_data["cursor_x"][i], raw_data["cursor_y"][i]]
        )
        track_ball_point_regualer = track_ball_point - center_point

        target_angel[i] = math.atan2(
            target_ball_point[1] - center_point[1],
            target_ball_point[0] - center_point[0],
        )

        track_angel[i] = math.atan2(
            track_ball_point[1] - center_point[1],
            track_ball_point[0] - center_point[0],
        )

        track_noise[i] = (
                distance_between_two_points((0, 0), track_ball_point_regualer)
                - radius_point
        )

    raw_data["pad_x"] = raw_data["cursor_x"]
    raw_data["pad_y"] = raw_data["cursor_y"]

    plt.plot(raw_data["pad_x"], label="x")
    plt.xlabel("Frame")
    plt.plot(raw_data["pad_y"], label="y")
    plt.ylabel("Value")

    plt.savefig("output/pad_x_y"+type)
    plt.xlim(220, 280)
    plt.ylim(100, 800)
    plt.savefig("output/pad_x_y_small"+type)
    plt.close()

    plt.plot(raw_data["pad_x"], raw_data["pad_y"])
    plt.xlabel("x")
    plt.ylabel("y")

    plt.savefig("output/pad_x_y_circle"+type)
    plt.close()

    track_noise = 10.0 * F.tanh(0.01 * track_noise)
    # 一阶差分法计算target和x的切割点
    # 异或求差分交集
    # data_index_target=torch.zeros(raw_data_len)
    # data_index_x=torch.zeros(raw_data_len)
    diffs_target = torch.diff(target_angel)
    # diffs_x=torch.diff(x_angel)

    start_point_target = torch.where(diffs_target < 0)[0] + 1
    # start_point_x = torch.where(diffs_x < 0)[0] + 1

    list_of_start_end = []
    list_of_target_angel = []
    list_of_track_angel = []
    list_of_error_angel = []
    list_of_track_noise = []
    list_of_data = []
    list_of_raw_x = []
    list_of_raw_y = []
    first_start_point = -1
    last_end_point = -1
    for i in range(1, len(start_point_target) - 1):
        # data_index_target[start_point_target[i]:start_point_target[i+1]-1] = i
        current_start_point = (
                track_angel[start_point_target[i]: start_point_target[i + 1] - 1].argmin()
                + start_point_target[i]
        ).item()
        if i == 1:
            first_start_point = current_start_point
        current_length = (
                track_angel[current_start_point: start_point_target[i + 1] - 1].argmax()
                - 1
        )
        current_end_point = (current_start_point + current_length).item()
        if i == len(start_point_target) - 2:
            last_end_point = current_end_point

        current_target_split = target_angel[current_start_point:current_end_point]
        current_track_split = track_angel[current_start_point:current_end_point]

        list_of_start_end.append([current_start_point, current_end_point])

        list_of_target_angel.append(current_target_split)
        list_of_track_angel.append(current_track_split)
        list_of_error_angel.append(current_track_split - current_target_split)
        list_of_track_noise.append(track_noise[current_start_point:current_end_point])
        list_of_raw_x.append(
            torch.tensor(raw_data["cursor_x"][current_start_point:current_end_point])
        )
        list_of_raw_y.append(
            torch.tensor(raw_data["cursor_y"][current_start_point:current_end_point])
        )

    entire_raw_x = (
            torch.tensor(
                raw_data["cursor_x"][:-1], dtype=torch.float
            )
            - center_point[0]
    )

    entire_raw_x = util.low_pass_filter(entire_raw_x, 0.5, 50)

    entire_raw_y = (
            torch.tensor(
                raw_data["cursor_y"][:-1], dtype=torch.float
            )
            - center_point[1]
    )

    entire_raw_y = util.low_pass_filter(entire_raw_y, 0.5, 50)

    entire_raw_vx = torch.diff(
        torch.tensor(
            raw_data["cursor_x"],
            dtype=torch.float,
        )
    )

    entire_raw_vx = util.low_pass_filter(entire_raw_vx, 0.4, 50)

    entire_raw_vy = torch.diff(
        torch.tensor(
            raw_data["cursor_y"],
            dtype=torch.float,
        )
    )

    entire_raw_vy = util.low_pass_filter(entire_raw_vy, 0.4, 50)

    raw_data_gx = torch.tensor(raw_data["target_x"][:-1], dtype=torch.float) - center_point[0]
    raw_data_gy = torch.tensor(raw_data["target_y"][:-1], dtype=torch.float) - center_point[1]

    # calc RT

    RT_r = torch.zeros(entire_raw_x.shape)

    angles_r = torch.atan2(entire_raw_y, entire_raw_x)
    angle_diffs_r = angles_r[1:] - angles_r[:-1]
    angle_diffs_r = (angle_diffs_r + np.pi) % (2 * np.pi) - np.pi

    cumulative_angles_r = torch.cumsum(angle_diffs_r, dim=0)
    entire_raw_r = torch.cat((torch.tensor([0.0]), cumulative_angles_r))

    angles_gr = torch.atan2(raw_data_gy, raw_data_gx)
    angle_diffs_gr = angles_gr[1:] - angles_gr[:-1]
    angle_diffs_gr = (angle_diffs_gr + np.pi) % (2 * np.pi) - np.pi

    cumulative_angles_gr = torch.cumsum(angle_diffs_gr, dim=0)
    entire_raw_gr = torch.cat((torch.tensor([0.0]), cumulative_angles_gr))

    entire_raw_r[torch.where(entire_raw_r < entire_raw_gr.min())] = entire_raw_gr.min()
    entire_raw_r[torch.where(entire_raw_r > entire_raw_gr.max())] = entire_raw_gr.max()

    for i in range(len(entire_raw_x) - 1):
        curve = entire_raw_gr - entire_raw_r[i]
        diff = curve[1:] * curve[:-1]
        zero_crossings = torch.where(diff <= 0)[0]
        RT_r[i] = i - zero_crossings[torch.argmin(torch.abs(zero_crossings - i))]

    # calc TE
    radius_xy = torch.sqrt(entire_raw_x * entire_raw_x + entire_raw_y * entire_raw_y)
    radius_gxgy = torch.sqrt(raw_data_gx * raw_data_gx + raw_data_gy * raw_data_gy)
    TE = radius_xy - radius_gxgy

    # calc PI
    point_diff_x = entire_raw_x[1:] - entire_raw_x[:-1]
    point_diff_y = entire_raw_y[1:] - entire_raw_y[:-1]
    point_diff = torch.sqrt(point_diff_x * point_diff_x + point_diff_y * point_diff_y)
    PI = torch.cat((torch.tensor([0.0]), point_diff))

    # raw_data["RT_r"] = RT_r.cpu().detach().numpy()
    # raw_data["TE"] = TE.cpu().detach().numpy()
    raw_data["filter_x"] = entire_raw_x.cpu().detach().numpy()
    raw_data["filter_y"] = entire_raw_y.cpu().detach().numpy()
    raw_data["filter_vx"] = entire_raw_vx.cpu().detach().numpy()
    raw_data["filter_vy"] = entire_raw_vy.cpu().detach().numpy()

    plt.plot(raw_data["filter_x"], label="x")
    plt.xlabel("Frame")
    plt.plot(raw_data["filter_y"], label="y")
    plt.ylabel("Value")

    plt.savefig("output/filter_x_y"+type)
    plt.xlim(220, 280)
    plt.ylim(100, 800)
    plt.savefig("output/filter_x_y_small"+type)
    plt.close()

    plt.plot(raw_data["filter_x"], raw_data["filter_y"])
    plt.xlabel("x")
    plt.ylabel("y")

    plt.savefig("output/filter_x_y_circle"+type)
    plt.close()

    return {
        "raw_data": raw_data,
        "start_end": list_of_start_end,
        "target_angel": list_of_target_angel,
        "track_angel": list_of_track_angel,
        "error_angel": list_of_error_angel,
        "track_noise": list_of_track_noise,
        "raw_x": list_of_raw_x,
        "raw_y": list_of_raw_y,
        "entire_raw_x": entire_raw_x,
        "entire_raw_y": entire_raw_y,
        "entire_raw_vx": entire_raw_vx,
        "entire_raw_vy": entire_raw_vy,
        "RT_r": RT_r,
        "TE": TE,
        "PI": PI
    }

def get_urtracker_data(path=r'./output/1726992875.266219.xlsx'):
    # x = torch.zeros([n, steps], dtype=torch.float64)
    # tobii on surface
    center_point = torch.tensor([432, 432])
    # center_point = torch.tensor([300, 300])
    radius_point = 252

    workbook = load_workbook(filename=path)

    sheet = workbook.active

    # 读取数据
    raw_data = {}
    keys = []  
    '''
    [time, target_x, target_y, cursor_x, cursor_y, Hex_x, Hex_y, Hex_z, 
    linear_x, linear_y, linear_z, pose_x, pose_y, pose_z, pose_rx, pose_ry, pose_rz]
    '''

    for row in sheet.iter_rows(values_only=True):
        # data.append(row)
        if len(raw_data.keys()) == 0:
            for key in row:
                if key is not None:
                    keys.append(key)
                    raw_data[key] = []
            continue

        for i in range(len(keys)):
            raw_data[keys[i]].append(row[i + 1])

    raw_data_ori = copy.deepcopy(raw_data)
    raw_data_len = len(raw_data["time"])
    target_angel = torch.zeros(raw_data_len)
    track_angel = torch.zeros(raw_data_len)
    track_noise = torch.zeros(raw_data_len)

    raw_x = torch.zeros(raw_data_len)
    raw_y = torch.zeros(raw_data_len)

    raw_data["ori_x"] = raw_data["cursor_x"]
    raw_data["ori_y"] = raw_data["cursor_y"]

    plt.plot(raw_data["ori_x"], label="x")
    plt.xlabel("Frame")
    plt.plot(raw_data["ori_y"], label="y")
    plt.ylabel("Value")

    plt.savefig("ori_x_y")
    plt.xlim(220, 280)
    plt.ylim(100, 800)
    plt.savefig("ori_x_y_small")
    plt.close()

    plt.plot(raw_data["ori_x"], raw_data["ori_y"])
    plt.xlabel("x")
    plt.ylabel("y")

    plt.savefig("ori_x_y_circle")
    plt.close()

    for i in range(raw_data_len):
        target_ball_point = torch.tensor(
            [raw_data["target_x"][i], raw_data["target_y"][i]]
        )
        target_ball_point_regualer = target_ball_point - center_point
        if raw_data["cursor_x"][i] is None:
            step = 0
            while raw_data["cursor_x"][i + step] is None and i + step + 1 < len(
                    raw_data["cursor_x"]
            ):
                step += 1

            if i == 0:
                # padding from next
                for j in range(i, i + step):
                    raw_data["cursor_x"][j] = raw_data["cursor_x"][i + step]
                    raw_data["cursor_y"][j] = raw_data["cursor_y"][i + step]

            elif i + step == len(raw_data["cursor_x"]):
                # padding from pre
                for j in range(i, i + step):
                    raw_data["cursor_x"][j] = raw_data["cursor_x"][i - 1]
                    raw_data["cursor_y"][j] = raw_data["cursor_y"][i - 1]
            else:
                # average padding
                for j in range(i, i + step):
                    raw_data["cursor_x"][j] = raw_data["cursor_x"][i - 1] + (
                            (j - i + 1) / (step + 1)
                    ) * (
                                                         raw_data["cursor_x"][i + step]
                                                         - raw_data["cursor_x"][i - 1]
                                                 )
                    raw_data["cursor_y"][j] = raw_data["cursor_y"][i - 1] + (
                            (j - i + 1) / (step + 1)
                    ) * (
                                                         raw_data["cursor_y"][i + step]
                                                         - raw_data["cursor_y"][i - 1]
                                                 )

        track_ball_point = torch.tensor(
            [raw_data["cursor_x"][i], raw_data["cursor_y"][i]]
        )
        track_ball_point_regualer = track_ball_point - center_point

        target_angel[i] = math.atan2(
            target_ball_point[1] - center_point[1],
            target_ball_point[0] - center_point[0],
        )

        track_angel[i] = math.atan2(
            track_ball_point[1] - center_point[1],
            track_ball_point[0] - center_point[0],
        )

        track_noise[i] = (
                distance_between_two_points((0, 0), track_ball_point_regualer)
                - radius_point
        )

    raw_data["pad_x"] = raw_data["cursor_x"]
    raw_data["pad_y"] = raw_data["cursor_y"]

    plt.plot(raw_data["pad_x"], label="x")
    plt.xlabel("Frame")
    plt.plot(raw_data["pad_y"], label="y")
    plt.ylabel("Value")

    plt.savefig("pad_x_y")
    plt.xlim(220, 280)
    plt.ylim(100, 800)
    plt.savefig("pad_x_y_small")
    plt.close()

    plt.plot(raw_data["pad_x"], raw_data["pad_y"])
    plt.xlabel("x")
    plt.ylabel("y")

    plt.savefig("pad_x_y_circle")
    plt.close()

    track_noise = 10.0 * F.tanh(0.01 * track_noise)
    # 一阶差分法计算target和x的切割点
    # 异或求差分交集
    # data_index_target=torch.zeros(raw_data_len)
    # data_index_x=torch.zeros(raw_data_len)
    diffs_target = torch.diff(target_angel)
    # diffs_x=torch.diff(x_angel)

    start_point_target = torch.where(diffs_target < 0)[0] + 1
    # start_point_x = torch.where(diffs_x < 0)[0] + 1

    list_of_start_end = []
    list_of_target_angel = []
    list_of_track_angel = []
    list_of_error_angel = []
    list_of_track_noise = []
    list_of_data = []
    list_of_raw_x = []
    list_of_raw_y = []
    first_start_point = -1
    last_end_point = -1
    for i in range(1, len(start_point_target) - 1):
        # data_index_target[start_point_target[i]:start_point_target[i+1]-1] = i
        current_start_point = (
                track_angel[start_point_target[i]: start_point_target[i + 1] - 1].argmin()
                + start_point_target[i]
        ).item()
        if i == 1:
            first_start_point = current_start_point
        current_length = (
                track_angel[current_start_point: start_point_target[i + 1] - 1].argmax()
                - 1
        )
        current_end_point = (current_start_point + current_length).item()
        if i == len(start_point_target) - 2:
            last_end_point = current_end_point

        current_target_split = target_angel[current_start_point:current_end_point]
        current_track_split = track_angel[current_start_point:current_end_point]

        list_of_start_end.append([current_start_point, current_end_point])

        list_of_target_angel.append(current_target_split)
        list_of_track_angel.append(current_track_split)
        list_of_error_angel.append(current_track_split - current_target_split)
        list_of_track_noise.append(track_noise[current_start_point:current_end_point])
        list_of_raw_x.append(
            torch.tensor(raw_data["cursor_x"][current_start_point:current_end_point])
        )
        list_of_raw_y.append(
            torch.tensor(raw_data["cursor_y"][current_start_point:current_end_point])
        )

    entire_raw_x = (
            torch.tensor(
                raw_data["cursor_x"][:-1], dtype=torch.float
            )
            - center_point[0]
    )

    entire_raw_x = util.low_pass_filter(entire_raw_x, 0.5, 50)

    entire_raw_y = (
            torch.tensor(
                raw_data["cursor_y"][:-1], dtype=torch.float
            )
            - center_point[1]
    )

    entire_raw_y = util.low_pass_filter(entire_raw_y, 0.5, 50)

    entire_raw_vx = torch.diff(
        torch.tensor(
            raw_data["cursor_x"],
            dtype=torch.float,
        )
    )

    entire_raw_vx = util.low_pass_filter(entire_raw_vx, 0.4, 50)

    entire_raw_vy = torch.diff(
        torch.tensor(
            raw_data["cursor_y"],
            dtype=torch.float,
        )
    )

    entire_raw_vy = util.low_pass_filter(entire_raw_vy, 0.4, 50)

    raw_data["filter_x"] = entire_raw_x.cpu().detach().numpy()
    raw_data["filter_y"] = entire_raw_y.cpu().detach().numpy()
    raw_data["filter_vx"] = entire_raw_vx.cpu().detach().numpy()
    raw_data["filter_vy"] = entire_raw_vy.cpu().detach().numpy()

    plt.plot(raw_data["filter_x"], label="x")
    plt.xlabel("Frame")
    plt.plot(raw_data["filter_y"], label="y")
    plt.ylabel("Value")

    plt.savefig("filter_x_y")
    plt.xlim(220, 280)
    plt.ylim(100, 800)
    plt.savefig("filter_x_y_small")
    plt.close()

    plt.plot(raw_data["filter_x"], raw_data["filter_y"])
    plt.xlabel("x")
    plt.ylabel("y")

    plt.savefig("filter_x_y_circle")
    plt.close()

    return {
        "raw_data": raw_data,
        "start_end": list_of_start_end,
        "target_angel": list_of_target_angel,
        "track_angel": list_of_track_angel,
        "error_angel": list_of_error_angel,
        "track_noise": list_of_track_noise,
        "raw_x": list_of_raw_x,
        "raw_y": list_of_raw_y,
        "entire_raw_x": entire_raw_x,
        "entire_raw_y": entire_raw_y,
        "entire_raw_vx": entire_raw_vx,
        "entire_raw_vy": entire_raw_vy,
    }


if __name__ == '__main__':
    get_urtracker_data(r'./output/1726992875.266219.xlsx')